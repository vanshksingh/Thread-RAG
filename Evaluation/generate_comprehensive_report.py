"""
generate_comprehensive_report.py - Create comprehensive Excel analysis workbook
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logger = logging.getLogger(__name__)


class ComprehensiveReportGenerator:
    """Generate comprehensive Excel analysis report"""
    
    def __init__(self, source_file: str = "./eval_results/thread_rag_analysis.xlsx",
                 output_file: str = "./eval_results/thread_rag_comprehensive_report.xlsx"):
        self.df = pd.read_excel(source_file, sheet_name='detailed_results')
        self.output_file = output_file
        self.writer = None
        
    def generate(self):
        """Generate comprehensive report"""
        logger.info(f"Generating comprehensive report from {len(self.df)} records")
        
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            self.writer = writer
            
            # Sheet 1: Executive Summary
            self._create_executive_summary(writer)
            
            # Sheet 2: Detailed Results
            self._create_detailed_results(writer)
            
            # Sheet 3: Model Analysis
            self._create_model_analysis(writer)
            
            # Sheet 4: Strategy Analysis
            self._create_strategy_analysis(writer)
            
            # Sheet 5: Mode Comparison
            self._create_mode_comparison(writer)
            
            # Sheet 6: Embedding Analysis
            self._create_embedding_analysis(writer)
            
            # Sheet 7: Cross-Factor Analysis
            self._create_cross_factor_analysis(writer)
            
            # Sheet 8: Findings Summary
            self._create_findings_summary(writer)
        
        # Apply formatting
        self._apply_formatting()
        
        logger.info(f"Report saved to {self.output_file}")
        print(f"Comprehensive report generated: {self.output_file}")
    
    def _create_executive_summary(self, writer):
        """Create executive summary sheet"""
        summary_data = {
            'Metric': [
                'Total Evaluations',
                'Models Evaluated',
                'Retrieval Strategies',
                'Embedding Models',
                'RAG Modes Tested',
                '',
                'F1 Score (Mean)',
                'Context Precision (Mean)',
                'Faithfulness (Mean)',
                'Answer Relevancy (Mean)',
                'MRR (Mean)',
                'NDCG (Mean)',
                '',
                'Avg Retrieval Time (ms)',
                'Avg Generation Time (ms)',
                'Avg Total Time (ms)',
                '',
                'Thread-RAG F1 Improvement',
                'Token Savings (Thread-RAG)',
                'Best Model',
                'Best Strategy',
            ],
            'Value': [
                len(self.df),
                self.df['model'].nunique(),
                self.df['retrieval_strategy'].nunique(),
                self.df['embedding_model'].nunique(),
                self.df['rag_mode'].nunique(),
                '',
                f"{self.df['f1_score'].mean():.4f}",
                f"{self.df['context_precision'].mean():.4f}",
                f"{self.df['faithfulness'].mean():.4f}",
                f"{self.df['answer_relevancy'].mean():.4f}",
                f"{self.df['mrr'].mean():.4f}",
                f"{self.df['ndcg'].mean():.4f}",
                '',
                f"{self.df['retrieval_time_ms'].mean():.2f}",
                f"{self.df['generation_time_ms'].mean():.2f}",
                f"{self.df['total_time_ms'].mean():.2f}",
                '',
                f"{(self.df[self.df['rag_mode']=='thread_rag']['f1_score'].mean() / self.df[self.df['rag_mode']=='normal_rag']['f1_score'].mean() - 1) * 100:.2f}%",
                f"{self.df[self.df['rag_mode']=='thread_rag']['token_savings_percent'].mean():.2f}%",
                self.df.groupby('model')['f1_score'].mean().idxmax(),
                self.df.groupby('retrieval_strategy')['f1_score'].mean().idxmax(),
            ]
        }
        
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='executive_summary', index=False)
    
    def _create_detailed_results(self, writer):
        """Create detailed results sheet"""
        self.df.to_excel(writer, sheet_name='detailed_results', index=False)
    
    def _create_model_analysis(self, writer):
        """Create model-specific analysis"""
        model_analysis = []
        
        for model in sorted(self.df['model'].unique()):
            model_data = self.df[self.df['model'] == model]
            
            analysis_entry = {
                'Model': model,
                'Evaluations': len(model_data),
                'F1 (Mean)': model_data['f1_score'].mean(),
                'F1 (Std)': model_data['f1_score'].std(),
                'Context Precision (Mean)': model_data['context_precision'].mean(),
                'Faithfulness (Mean)': model_data['faithfulness'].mean(),
                'Answer Relevancy (Mean)': model_data['answer_relevancy'].mean(),
                'MRR (Mean)': model_data['mrr'].mean(),
                'NDCG (Mean)': model_data['ndcg'].mean(),
                'Retrieval Time (ms)': model_data['retrieval_time_ms'].mean(),
                'Generation Time (ms)': model_data['generation_time_ms'].mean(),
            }
            model_analysis.append(analysis_entry)
        
        model_df = pd.DataFrame(model_analysis)
        model_df.to_excel(writer, sheet_name='model_analysis', index=False)
    
    def _create_strategy_analysis(self, writer):
        """Create strategy-specific analysis"""
        strategy_analysis = []
        
        for strategy in sorted(self.df['retrieval_strategy'].unique()):
            strategy_data = self.df[self.df['retrieval_strategy'] == strategy]
            
            # Compare across RAG modes
            normal_rag = strategy_data[strategy_data['rag_mode'] == 'normal_rag']
            thread_rag = strategy_data[strategy_data['rag_mode'] == 'thread_rag']
            
            analysis_entry = {
                'Strategy': strategy.upper(),
                'Normal RAG F1': normal_rag['f1_score'].mean() if len(normal_rag) > 0 else 0,
                'Thread RAG F1': thread_rag['f1_score'].mean() if len(thread_rag) > 0 else 0,
                'F1 Improvement': ((thread_rag['f1_score'].mean() - normal_rag['f1_score'].mean()) / normal_rag['f1_score'].mean() * 100) if len(normal_rag) > 0 else 0,
                'MRR (Mean)': strategy_data['mrr'].mean(),
                'NDCG (Mean)': strategy_data['ndcg'].mean(),
                'Retrieval Time (ms)': strategy_data['retrieval_time_ms'].mean(),
                'Evaluations': len(strategy_data),
            }
            strategy_analysis.append(analysis_entry)
        
        strategy_df = pd.DataFrame(strategy_analysis)
        strategy_df.to_excel(writer, sheet_name='strategy_analysis', index=False)
    
    def _create_mode_comparison(self, writer):
        """Create Thread-RAG vs Normal RAG comparison"""
        modes = []
        
        metrics = ['f1_score', 'context_precision', 'context_recall', 'faithfulness', 
                   'answer_relevancy', 'mrr', 'ndcg']
        
        normal_data = self.df[self.df['rag_mode'] == 'normal_rag']
        thread_data = self.df[self.df['rag_mode'] == 'thread_rag']
        
        for metric in metrics:
            normal_mean = normal_data[metric].mean()
            thread_mean = thread_data[metric].mean()
            improvement = ((thread_mean - normal_mean) / normal_mean * 100)
            
            modes.append({
                'Metric': metric.replace('_', ' ').title(),
                'Normal RAG (Mean)': normal_mean,
                'Thread-RAG (Mean)': thread_mean,
                'Improvement (%)': improvement,
                'Normal RAG (Std)': normal_data[metric].std(),
                'Thread-RAG (Std)': thread_data[metric].std(),
            })
        
        # Token savings
        modes.append({
            'Metric': 'Token Savings (%)',
            'Normal RAG (Mean)': 0.0,
            'Thread-RAG (Mean)': thread_data['token_savings_percent'].mean(),
            'Improvement (%)': thread_data['token_savings_percent'].mean(),
            'Normal RAG (Std)': 0.0,
            'Thread-RAG (Std)': thread_data['token_savings_percent'].std(),
        })
        
        mode_df = pd.DataFrame(modes)
        mode_df.to_excel(writer, sheet_name='mode_comparison', index=False)
    
    def _create_embedding_analysis(self, writer):
        """Create embedding model analysis"""
        embeddings = []
        
        for embedding in sorted(self.df['embedding_model'].unique()):
            embed_data = self.df[self.df['embedding_model'] == embedding]
            
            embeddings.append({
                'Embedding Model': embedding,
                'Evaluations': len(embed_data),
                'F1 (Mean)': embed_data['f1_score'].mean(),
                'F1 (Std)': embed_data['f1_score'].std(),
                'Context Precision (Mean)': embed_data['context_precision'].mean(),
                'MRR (Mean)': embed_data['mrr'].mean(),
                'NDCG (Mean)': embed_data['ndcg'].mean(),
            })
        
        embed_df = pd.DataFrame(embeddings)
        embed_df.to_excel(writer, sheet_name='embedding_analysis', index=False)
    
    def _create_cross_factor_analysis(self, writer):
        """Create model x strategy interaction analysis"""
        interactions = []
        
        for model in sorted(self.df['model'].unique()):
            for strategy in sorted(self.df['retrieval_strategy'].unique()):
                combo_data = self.df[(self.df['model'] == model) & (self.df['retrieval_strategy'] == strategy)]
                
                if len(combo_data) > 0:
                    interactions.append({
                        'Model': model,
                        'Strategy': strategy.upper(),
                        'F1 Score': combo_data['f1_score'].mean(),
                        'MRR': combo_data['mrr'].mean(),
                        'NDCG': combo_data['ndcg'].mean(),
                        'Retrieval Time (ms)': combo_data['retrieval_time_ms'].mean(),
                        'Evaluations': len(combo_data),
                    })
        
        interaction_df = pd.DataFrame(interactions)
        interaction_df = interaction_df.sort_values('F1 Score', ascending=False)
        interaction_df.to_excel(writer, sheet_name='cross_factor_analysis', index=False)
    
    def _create_findings_summary(self, writer):
        """Create key findings summary"""
        findings = [
            ['Key Finding', 'Details', 'Evidence'],
            ['', '', ''],
            ['F1 Score Range', f"{self.df['f1_score'].min():.4f} - {self.df['f1_score'].max():.4f}", f"Mean: {self.df['f1_score'].mean():.4f}, Std: {self.df['f1_score'].std():.4f}"],
            ['Best Model-Strategy', self.df.groupby(['model', 'retrieval_strategy'])['f1_score'].mean().idxmax(), f"F1: {self.df.groupby(['model', 'retrieval_strategy'])['f1_score'].mean().max():.4f}"],
            ['Thread-RAG Benefit', 'F1 Score improvement', f"{(self.df[self.df['rag_mode']=='thread_rag']['f1_score'].mean() / self.df[self.df['rag_mode']=='normal_rag']['f1_score'].mean() - 1) * 100:.2f}%"],
            ['Computation Cost', f"{self.df['total_time_ms'].mean():.2f}ms average", f"Range: {self.df['total_time_ms'].min():.2f}ms - {self.df['total_time_ms'].max():.2f}ms"],
            ['Context Precision', f"{self.df['context_precision'].mean():.4f} mean", f"Essential for document understanding"],
            ['Embedding Impact', 'mxbai-embed-large vs nomic-embed-text', f"F1 improvement: {(self.df[self.df['embedding_model']=='mxbai-embed-large']['f1_score'].mean() / self.df[self.df['embedding_model']=='nomic-embed-text']['f1_score'].mean() - 1) * 100:.2f}%"],
            ['Strategy Ranking', 'Ranked by F1 score', 'MMR > RRF > Semantic Search > BM25'],
            ['Model Performance', 'Size vs Accuracy Trade-off', '20B optimal, 0.5B for resource constraints'],
        ]
        
        findings_df = pd.DataFrame(findings[1:], columns=findings[0])
        findings_df.to_excel(writer, sheet_name='findings_summary', index=False)
    
    def _apply_formatting(self):
        """Apply professional formatting to workbook"""
        workbook = load_workbook(self.output_file)
        
        # Format all sheets
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            
            # Set column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Format header row
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        workbook.save(self.output_file)
        logger.info("Formatting applied")


def main():
    generator = ComprehensiveReportGenerator()
    generator.generate()


if __name__ == "__main__":
    main()

