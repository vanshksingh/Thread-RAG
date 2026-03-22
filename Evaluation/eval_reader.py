"""
eval_reader.py - Read, analyze, and visualize evaluation results
"""

import logging
from typing import Dict, List, Optional, Tuple
from pathlib import Path
from datetime import datetime
import openpyxl
from collections import defaultdict
import statistics

from eval_config import EVAL_RESULTS_FILE, RAG_MODES, RETRIEVAL_STRATEGIES

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EvalReader:
    """Reads and analyzes evaluation results from Excel"""
    
    def __init__(self, excel_file: str = EVAL_RESULTS_FILE):
        self.excel_file = excel_file
        self.data = None
        self.wb = None
        self._load_workbook()

    def _load_workbook(self):
        """Load Excel workbook"""
        if not Path(self.excel_file).exists():
            logger.error(f"Excel file not found: {self.excel_file}")
            return False
        
        try:
            self.wb = openpyxl.load_workbook(self.excel_file, data_only=True)
            logger.info(f"Loaded Excel file: {self.excel_file}")
            return True
        except Exception as e:
            logger.error(f"Failed to load Excel: {e}")
            return False

    def get_detailed_results(self, filters: Optional[Dict] = None) -> List[Dict]:
        """Get detailed results with optional filtering"""
        if "detailed" not in self.wb.sheetnames:
            logger.error("'detailed' sheet not found")
            return []
        
        ws = self.wb["detailed"]
        headers = [cell.value for cell in ws[1]]
        results = []
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            result_dict = {header: value for header, value in zip(headers, row)}
            
            # Apply filters
            if filters:
                skip = False
                for key, value in filters.items():
                    if key in result_dict and result_dict[key] != value:
                        skip = True
                        break
                if skip:
                    continue
            
            results.append(result_dict)
        
        return results

    def get_summary_statistics(self) -> Dict:
        """Get summary statistics from summary sheet"""
        if "summary" not in self.wb.sheetnames:
            logger.error("'summary' sheet not found")
            return {}
        
        ws = self.wb["summary"]
        headers = [cell.value for cell in ws[1]]
        
        # Get last (most recent) row
        if ws.max_row <= 1:
            return {}
        
        last_row = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
        
        summary = {header: value for header, value in zip(headers, last_row)}
        return summary

    def compare_rag_modes(self) -> Dict[str, Dict]:
        """Compare metrics across RAG modes"""
        results = self.get_detailed_results({"Status": "completed"})
        
        comparison = {}
        
        for mode in RAG_MODES.keys():
            mode_results = [r for r in results if r.get("RAG_Mode") == mode]
            
            if not mode_results:
                continue
            
            metrics = {
                "faithfulness": [r.get("Faithfulness", 0) for r in mode_results],
                "answer_relevancy": [r.get("Answer_Relevancy", 0) for r in mode_results],
                "context_precision": [r.get("Context_Precision", 0) for r in mode_results],
                "context_recall": [r.get("Context_Recall", 0) for r in mode_results],
                "mrr": [r.get("MRR", 0) for r in mode_results],
                "ndcg": [r.get("NDCG", 0) for r in mode_results],
                "f1_score": [r.get("F1_Score", 0) for r in mode_results],
                "perplexity": [r.get("Perplexity", 0) for r in mode_results],
            }
            
            comparison[mode] = {
                "count": len(mode_results),
                **{
                    metric: {
                        "mean": statistics.mean(values) if values else 0,
                        "median": statistics.median(values) if values else 0,
                        "std": statistics.stdev(values) if len(values) > 1 else 0,
                        "min": min(values) if values else 0,
                        "max": max(values) if values else 0
                    }
                    for metric, values in metrics.items()
                }
            }
        
        return comparison

    def compare_retrieval_strategies(self) -> Dict[str, Dict]:
        """Compare metrics across retrieval strategies"""
        results = self.get_detailed_results({"Status": "completed"})
        
        comparison = {}
        
        for strategy in RETRIEVAL_STRATEGIES.keys():
            strategy_results = [r for r in results if r.get("Retrieval_Strategy") == strategy]
            
            if not strategy_results:
                continue
            
            metrics = {
                "mrr": [r.get("MRR", 0) for r in strategy_results],
                "ndcg": [r.get("NDCG", 0) for r in strategy_results],
                "f1_score": [r.get("F1_Score", 0) for r in strategy_results],
                "retrieval_time_ms": [r.get("Retrieval_Time_ms", 0) for r in strategy_results],
                "total_time_ms": [r.get("Total_Time_ms", 0) for r in strategy_results],
            }
            
            comparison[strategy] = {
                "count": len(strategy_results),
                **{
                    metric: {
                        "mean": statistics.mean(values) if values else 0,
                        "median": statistics.median(values) if values else 0,
                        "std": statistics.stdev(values) if len(values) > 1 else 0,
                        "min": min(values) if values else 0,
                        "max": max(values) if values else 0
                    }
                    for metric, values in metrics.items()
                }
            }
        
        return comparison

    def compare_models(self) -> Dict[str, Dict]:
        """Compare metrics across main models"""
        results = self.get_detailed_results({"Status": "completed"})
        
        comparison = {}
        
        # Get unique models
        models = set(r.get("Main_Model") for r in results if r.get("Main_Model"))
        
        for model in sorted(models):
            model_results = [r for r in results if r.get("Main_Model") == model]
            
            metrics = {
                "answer_relevancy": [r.get("Answer_Relevancy", 0) for r in model_results],
                "f1_score": [r.get("F1_Score", 0) for r in model_results],
                "generation_time_ms": [r.get("Generation_Time_ms", 0) for r in model_results],
                "perplexity": [r.get("Perplexity", 0) for r in model_results],
            }
            
            comparison[model] = {
                "count": len(model_results),
                **{
                    metric: {
                        "mean": statistics.mean(values) if values else 0,
                        "median": statistics.median(values) if values else 0,
                        "std": statistics.stdev(values) if len(values) > 1 else 0,
                    }
                    for metric, values in metrics.items()
                }
            }
        
        return comparison

    def compare_embedding_models(self) -> Dict[str, Dict]:
        """Compare metrics across embedding models"""
        results = self.get_detailed_results({"Status": "completed"})
        
        comparison = {}
        
        # Get unique embedding models
        emb_models = set(r.get("Embedding_Model") for r in results if r.get("Embedding_Model"))
        
        for emb_model in sorted(emb_models):
            emb_results = [r for r in results if r.get("Embedding_Model") == emb_model]
            
            metrics = {
                "context_precision": [r.get("Context_Precision", 0) for r in emb_results],
                "context_recall": [r.get("Context_Recall", 0) for r in emb_results],
                "mrr": [r.get("MRR", 0) for r in emb_results],
                "ndcg": [r.get("NDCG", 0) for r in emb_results],
                "retrieval_time_ms": [r.get("Retrieval_Time_ms", 0) for r in emb_results],
            }
            
            comparison[emb_model] = {
                "count": len(emb_results),
                **{
                    metric: {
                        "mean": statistics.mean(values) if values else 0,
                        "median": statistics.median(values) if values else 0,
                        "std": statistics.stdev(values) if len(values) > 1 else 0,
                    }
                    for metric, values in metrics.items()
                }
            }
        
        return comparison

    def get_failures(self) -> List[Dict]:
        """Get all failed evaluations"""
        return self.get_detailed_results({"Status": "failed"})

    def print_comparison_report(self):
        """Print comprehensive comparison report"""
        logger.info("\n" + "="*80)
        logger.info("EVALUATION COMPARISON REPORT")
        logger.info("="*80)
        
        # Summary
        summary = self.get_summary_statistics()
        if summary:
            logger.info("\n📊 SUMMARY (Latest Run):")
            logger.info(f"  Test Run ID: {summary.get('Test_Run_ID', 'N/A')}")
            logger.info(f"  Total Questions: {summary.get('Total_Questions', 0)}")
            logger.info(f"  Completed: {summary.get('Completed', 0)}")
            logger.info(f"  Failed: {summary.get('Failed', 0)}")
            logger.info(f"  Avg Faithfulness: {summary.get('Average_Faithfulness', 0):.4f}")
            logger.info(f"  Avg Answer Relevancy: {summary.get('Average_Answer_Relevancy', 0):.4f}")
            logger.info(f"  Avg MRR: {summary.get('Average_MRR', 0):.4f}")
            logger.info(f"  Avg NDCG: {summary.get('Average_NDCG', 0):.4f}")
        
        # RAG Mode Comparison
        mode_comparison = self.compare_rag_modes()
        if mode_comparison:
            logger.info("\n🎯 RAG MODE COMPARISON:")
            for mode, stats in mode_comparison.items():
                logger.info(f"\n  {mode.upper()} (n={stats['count']}):")
                for metric in ["faithfulness", "answer_relevancy", "mrr", "ndcg", "f1_score"]:
                    if metric in stats:
                        m = stats[metric]
                        logger.info(
                            f"    {metric:20s}: "
                            f"mean={m['mean']:.4f}, median={m['median']:.4f}, "
                            f"std={m['std']:.4f}"
                        )
        
        # Strategy Comparison
        strategy_comparison = self.compare_retrieval_strategies()
        if strategy_comparison:
            logger.info("\n🔍 RETRIEVAL STRATEGY COMPARISON:")
            for strategy, stats in strategy_comparison.items():
                logger.info(f"\n  {strategy.upper()} (n={stats['count']}):")
                for metric in ["mrr", "ndcg", "retrieval_time_ms"]:
                    if metric in stats:
                        m = stats[metric]
                        logger.info(
                            f"    {metric:20s}: "
                            f"mean={m['mean']:.4f}, median={m['median']:.4f}, "
                            f"std={m['std']:.4f}"
                        )
        
        # Model Comparison
        model_comparison = self.compare_models()
        if model_comparison:
            logger.info("\n🤖 MODEL COMPARISON:")
            for model, stats in model_comparison.items():
                logger.info(f"\n  {model} (n={stats['count']}):")
                for metric in ["answer_relevancy", "f1_score", "generation_time_ms"]:
                    if metric in stats:
                        m = stats[metric]
                        logger.info(
                            f"    {metric:20s}: "
                            f"mean={m['mean']:.4f}, std={m['std']:.4f}"
                        )
        
        # Failures
        failures = self.get_failures()
        if failures:
            logger.warning(f"\n⚠️  FAILURES ({len(failures)} total):")
            for failure in failures[:5]:
                logger.warning(f"  Q: {failure.get('Question', '')[:50]}...")
                logger.warning(f"     Error: {failure.get('Error_Message', 'Unknown')}")
        
        logger.info("\n" + "="*80 + "\n")

    def export_comparison_to_text(self, output_file: str = "eval_comparison.txt"):
        """Export comparison to text file"""
        with open(output_file, 'w') as f:
            f.write("="*80 + "\n")
            f.write("EVALUATION COMPARISON REPORT\n")
            f.write("="*80 + "\n\n")
            
            # Summary
            summary = self.get_summary_statistics()
            if summary:
                f.write("SUMMARY (Latest Run):\n")
                f.write(f"Test Run ID: {summary.get('Test_Run_ID', 'N/A')}\n")
                f.write(f"Total Questions: {summary.get('Total_Questions', 0)}\n")
                f.write(f"Completed: {summary.get('Completed', 0)}\n")
                f.write(f"Failed: {summary.get('Failed', 0)}\n\n")
            
            # Mode Comparison
            mode_comparison = self.compare_rag_modes()
            if mode_comparison:
                f.write("RAG MODE COMPARISON:\n")
                f.write("-" * 80 + "\n")
                for mode, stats in mode_comparison.items():
                    f.write(f"\n{mode.upper()} (n={stats['count']}):\n")
                    for metric in ["faithfulness", "answer_relevancy", "mrr", "ndcg"]:
                        if metric in stats:
                            m = stats[metric]
                            f.write(
                                f"  {metric:20s}: mean={m['mean']:.4f}, "
                                f"median={m['median']:.4f}, std={m['std']:.4f}\n"
                            )
            
            # Strategy Comparison
            strategy_comparison = self.compare_retrieval_strategies()
            if strategy_comparison:
                f.write("\n\nRETRIEVAL STRATEGY COMPARISON:\n")
                f.write("-" * 80 + "\n")
                for strategy, stats in strategy_comparison.items():
                    f.write(f"\n{strategy.upper()} (n={stats['count']}):\n")
                    for metric in ["mrr", "ndcg", "retrieval_time_ms"]:
                        if metric in stats:
                            m = stats[metric]
                            f.write(
                                f"  {metric:20s}: mean={m['mean']:.4f}, "
                                f"median={m['median']:.4f}\n"
                            )
        
        logger.info(f"Comparison exported to {output_file}")

    def run_specific_query(self, query_filters: Dict) -> List[Dict]:
        """Run specific query/filter on results
        
        Example filters:
        {
            "Main_Model": "gpt-oss:20b",
            "RAG_Mode": "thread_rag",
            "Retrieval_Strategy": "rrf"
        }
        """
        results = self.get_detailed_results(query_filters)
        logger.info(f"Query returned {len(results)} results")
        return results

    def get_best_configuration(self, metric: str = "f1_score", top_k: int = 10) -> List[Dict]:
        """Get top-K best performing configurations"""
        results = self.get_detailed_results({"Status": "completed"})
        
        if not results or metric not in results[0]:
            logger.warning(f"Metric '{metric}' not found in results")
            return []
        
        # Sort by metric descending
        sorted_results = sorted(
            results,
            key=lambda x: x.get(metric, 0),
            reverse=True
        )
        
        return sorted_results[:top_k]


if __name__ == "__main__":
    reader = EvalReader()
    reader.print_comparison_report()
    reader.export_comparison_to_text()
    
    # Example: Query specific configuration
    logger.info("\n\nExample Query Results:")
    results = reader.run_specific_query({
        "RAG_Mode": "thread_rag",
        "Retrieval_Strategy": "rrf"
    })
    logger.info(f"Found {len(results)} results for thread_rag + rrf")
    
    # Get best configuration
    best = reader.get_best_configuration("f1_score", top_k=5)
    if best:
        logger.info("\nTop 5 Configurations by F1 Score:")
        for i, config in enumerate(best, 1):
            logger.info(
                f"{i}. {config.get('Main_Model')} + "
                f"{config.get('Embedding_Model')} + "
                f"{config.get('RAG_Mode')} + "
                f"{config.get('Retrieval_Strategy')}: {config.get('F1_Score', 0):.4f}"
            )

