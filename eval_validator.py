"""
eval_validator.py - Validation and Quality Assurance for Evaluation Data
"""

import logging
from typing import List
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("Warning: openpyxl not installed. Run: pip install openpyxl")

from eval_config import EVAL_RESULTS_FILE, EXCEL_SHEETS

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class EvalValidator:
    """Validates evaluation results for data integrity and consistency"""
    
    def __init__(self, excel_file: str = EVAL_RESULTS_FILE):
        self.excel_file = excel_file
        self.issues = []
        self.warnings = []
        self.summary = {}

    def validate_all(self) -> bool:
        """Run all validations"""
        logger.info("Starting comprehensive validation...")
        
        if not Path(self.excel_file).exists():
            logger.error(f"Excel file not found: {self.excel_file}")
            return False

        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            # Validate structure
            self._validate_sheets(wb)
            self._validate_headers(wb)
            
            # Validate data
            self._validate_detailed_sheet(wb)
            self._validate_summary_sheet(wb)
            self._validate_data_types(wb)
            self._validate_data_ranges(wb)
            self._validate_consistency(wb)
            
            # Report
            self._report()
            
            return len(self.issues) == 0
        
        except Exception as e:
            logger.error(f"Validation failed: {e}")
            return False

    def _validate_sheets(self, wb):
        """Check all required sheets exist"""
        expected_sheets = list(EXCEL_SHEETS.keys())
        actual_sheets = wb.sheetnames
        
        missing = set(expected_sheets) - set(actual_sheets)
        if missing:
            self.issues.append(f"Missing sheets: {missing}")
        
        extra = set(actual_sheets) - set(expected_sheets)
        if extra:
            self.warnings.append(f"Extra sheets found: {extra}")

    def _validate_headers(self, wb):
        """Check headers match expected columns"""
        for sheet_name, config in EXCEL_SHEETS.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            expected_headers = config["columns"]
            actual_headers = [cell.value for cell in ws[1]]
            
            if actual_headers != expected_headers:
                self.warnings.append(
                    f"Sheet '{sheet_name}': Headers mismatch. "
                    f"Expected {expected_headers}, got {actual_headers}"
                )

    def _validate_detailed_sheet(self, wb):
        """Validate detailed results sheet"""
        if "detailed" not in wb.sheetnames:
            return
        
        ws = wb["detailed"]
        logger.info(f"Validating {ws.max_row - 1} rows in detailed sheet...")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), 2):
            # Row ID should be numeric
            if row[0].value is not None:
                try:
                    int(row[0].value)
                except:
                    self.issues.append(f"Row {row_idx}: Invalid Row_ID (should be numeric)")

            # Timestamp should be valid ISO format
            if row[1].value:
                try:
                    datetime.fromisoformat(row[1].value)
                except:
                    self.warnings.append(f"Row {row_idx}: Invalid timestamp format")

            # Status should be "completed" or "failed"
            status = row[21].value
            if status and status not in ["completed", "failed", "pending"]:
                self.issues.append(f"Row {row_idx}: Invalid status '{status}'")

            # Numeric fields should be numbers
            numeric_cols = {
                "retrieved_chunks": 8,
                "faithfulness": 10,
                "mrr": 15,
                "ndcg": 16,
                "f1_score": 17,
                "retrieval_time_ms": 18,
                "generation_time_ms": 19,
                "total_time_ms": 20
            }
            
            for field, col in numeric_cols.items():
                value = row[col].value
                if value is not None:
                    try:
                        float(value)
                    except:
                        if str(value).strip():  # Skip empty cells
                            self.warnings.append(
                                f"Row {row_idx}: Non-numeric value in '{field}': {value}"
                            )

    def _validate_summary_sheet(self, wb):
        """Validate summary sheet"""
        if "summary" not in wb.sheetnames:
            return
        
        ws = wb["summary"]
        logger.info(f"Validating {ws.max_row - 1} summary entries...")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), 2):
            # Check numeric columns
            numeric_cols = range(5, 12)  # Total_Questions through Average_NDCG
            for col_idx in numeric_cols:
                value = row[col_idx].value
                if value is not None:
                    try:
                        float(value)
                    except:
                        self.warnings.append(
                            f"Summary row {row_idx}: Non-numeric in column {col_idx}"
                        )

    def _validate_data_types(self, wb):
        """Validate data type consistency"""
        if "detailed" not in wb.sheetnames:
            return
        
        ws = wb["detailed"]
        
        # Track unique models to ensure consistency
        models = set()
        embedding_models = set()
        rag_modes = set()
        strategies = set()
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[6]:  # Main_Model
                models.add(row[6])
            if row[7]:  # Embedding_Model
                embedding_models.add(row[7])
            if row[4]:  # RAG_Mode
                rag_modes.add(row[4])
            if row[5]:  # Retrieval_Strategy
                strategies.add(row[5])
        
        self.summary["unique_models"] = len(models)
        self.summary["unique_embedding_models"] = len(embedding_models)
        self.summary["unique_rag_modes"] = len(rag_modes)
        self.summary["unique_strategies"] = len(strategies)

    def _validate_data_ranges(self, wb):
        """Validate data value ranges"""
        if "detailed" not in wb.sheetnames:
            return
        
        ws = wb["detailed"]
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), 2):
            # Faithfulness should be 0-1
            if row[10] is not None:
                try:
                    val = float(row[10])
                    if not (0 <= val <= 1):
                        self.warnings.append(
                            f"Row {row_idx}: Faithfulness {val} out of range [0-1]"
                        )
                except:
                    pass

            # F1 Score should be 0-1
            if row[17] is not None:
                try:
                    val = float(row[17])
                    if not (0 <= val <= 1):
                        self.warnings.append(
                            f"Row {row_idx}: F1 Score {val} out of range [0-1]"
                        )
                except:
                    pass

            # Times should be positive
            for col, field in [(18, "retrieval_time"), (19, "generation_time"), (20, "total_time")]:
                if row[col] is not None:
                    try:
                        val = float(row[col])
                        if val < 0:
                            self.issues.append(
                                f"Row {row_idx}: Negative {field}: {val}"
                            )
                    except:
                        pass

    def _validate_consistency(self, wb):
        """Check cross-sheet consistency"""
        if "detailed" not in wb.sheetnames or "summary" not in wb.sheetnames:
            return
        
        ws_detailed = wb["detailed"]
        ws_summary = wb["summary"]
        
        # Count completed rows in detailed
        completed_count = 0
        failed_count = 0
        
        for row in ws_detailed.iter_rows(min_row=2, max_row=ws_detailed.max_row, values_only=True):
            status = row[21]
            if status == "completed":
                completed_count += 1
            elif status == "failed":
                failed_count += 1
        
        # Check against summary (last row should have these numbers)
        last_summary_row = ws_summary.max_row
        if last_summary_row > 1:
            summary_total = ws_summary[f"C{last_summary_row}"].value
            summary_completed = ws_summary[f"D{last_summary_row}"].value
            summary_failed = ws_summary[f"E{last_summary_row}"].value
            
            if summary_total and summary_total != ws_detailed.max_row - 1:
                self.warnings.append(
                    f"Summary reports {summary_total} rows, but detailed has {ws_detailed.max_row - 1}"
                )
            
            if summary_completed and summary_completed != completed_count:
                self.warnings.append(
                    f"Summary reports {summary_completed} completed, but detailed has {completed_count}"
                )

    def _report(self):
        """Print validation report"""
        logger.info("\n" + "="*60)
        logger.info("VALIDATION REPORT")
        logger.info("="*60)
        
        if self.issues:
            logger.error(f"\n❌ {len(self.issues)} CRITICAL ISSUES:")
            for issue in self.issues:
                logger.error(f"  - {issue}")
        else:
            logger.info("\n✅ No critical issues found")
        
        if self.warnings:
            logger.warning(f"\n⚠️  {len(self.warnings)} WARNINGS:")
            for warning in self.warnings[:10]:  # Show first 10
                logger.warning(f"  - {warning}")
            if len(self.warnings) > 10:
                logger.warning(f"  ... and {len(self.warnings) - 10} more")
        
        if self.summary:
            logger.info(f"\n📊 SUMMARY STATISTICS:")
            for key, value in self.summary.items():
                logger.info(f"  - {key}: {value}")
        
        logger.info("="*60 + "\n")

    def get_issues(self) -> List[str]:
        """Get list of all issues"""
        return self.issues

    def get_warnings(self) -> List[str]:
        """Get list of all warnings"""
        return self.warnings


if __name__ == "__main__":
    validator = EvalValidator()
    is_valid = validator.validate_all()
    exit(0 if is_valid else 1)


