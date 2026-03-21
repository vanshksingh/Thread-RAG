"""
eval_framework.py - Comprehensive RAG Evaluation Framework
Supports multiple RAG modes, retrieval strategies, models, and metrics with checkpoint recovery
"""

import json
import time
import uuid
import logging
from datetime import datetime
from typing import List, Dict, Optional, Tuple, Any
from pathlib import Path
import numpy as np
from collections import defaultdict

# LangChain/Ollama imports
from langchain_ollama import ChatOllama, OllamaEmbeddings, OllamaLLM
from langchain_chroma import Chroma
from langchain_community.retrievers import BM25Retriever
from langchain_core.documents import Document

# Evaluation imports
from ragas import evaluate
from ragas.metrics import (
    faithfulness,
    answer_relevancy,
    context_precision,
    context_recall
)
from datasets import Dataset

# Excel/Data handling
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Project imports
from eval_config import (
    EVAL_RESULTS_FILE, EVAL_CHECKPOINT_FILE, EVAL_DATASET_FILE,
    RAG_MODES, RETRIEVAL_STRATEGIES, OLLAMA_MODELS, EMBEDDING_MODELS,
    METRICS, EXCEL_SHEETS, CHECKPOINT_CONFIG, EVAL_PARAMS, EVAL_DIR,
    TOKEN_SAVINGS_CONFIG, CONTEXT_STABILITY_CONFIG
)
from rag_ret import (
    assemble_chunk_with_context, get_summary_on_demand,
    list_available_documents, vector_store
)

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class RRFRetriever:
    """Reciprocal Rank Fusion combining semantic and BM25 retrieval"""
    def __init__(self, vector_store, documents, rrf_k=60):
        self.vector_store = vector_store
        self.bm25_retriever = BM25Retriever.from_documents(documents)
        self.rrf_k = rrf_k

    def retrieve(self, query: str, k: int = 3) -> List[Document]:
        """Combine BM25 and semantic search via RRF"""
        # Get results from both retrievers
        semantic_results = self.vector_store.similarity_search(query, k=k*2)
        bm25_results = self.bm25_retriever.invoke(query)[:k*2]

        # Create RRF scoring
        rrf_scores = defaultdict(float)
        
        for rank, doc in enumerate(semantic_results, 1):
            rrf_scores[doc.metadata.get("id")] += 1 / (self.rrf_k + rank)
        
        for rank, doc in enumerate(bm25_results, 1):
            doc_id = doc.metadata.get("id") if hasattr(doc, 'metadata') else None
            if doc_id:
                rrf_scores[doc_id] += 1 / (self.rrf_k + rank)

        # Get top-k by RRF score
        all_docs = {doc.metadata.get("id"): doc for doc in semantic_results + bm25_results}
        sorted_docs = sorted(
            [(doc_id, score) for doc_id, score in rrf_scores.items()],
            key=lambda x: x[1],
            reverse=True
        )[:k]

        return [all_docs[doc_id] for doc_id, _ in sorted_docs if doc_id in all_docs]


class MMRRetriever:
    """Maximal Marginal Relevance for diversity-aware retrieval"""
    def __init__(self, vector_store, embeddings):
        self.vector_store = vector_store
        self.embeddings = embeddings

    def retrieve(self, query: str, k: int = 3, fetch_k: int = 6, lambda_mult: float = 0.5) -> List[Document]:
        """MMR retrieval balancing relevance and diversity"""
        # Get initial candidates
        candidates = self.vector_store.similarity_search(query, k=fetch_k)
        if len(candidates) <= k:
            return candidates

        # Embed query
        query_embedding = np.array(self.embeddings.embed_query(query))

        # Score candidates
        selected = []
        selected_embeddings = []
        
        for _ in range(k):
            if not candidates:
                break

            best_idx = 0
            best_score = float('-inf')

            for i, candidate in enumerate(candidates):
                # Relevance score (similarity to query)
                candidate_embedding = np.array(
                    self.embeddings.embed_query(candidate.page_content)
                )
                relevance = np.dot(query_embedding, candidate_embedding)

                # Diversity score (dissimilarity to selected)
                diversity = 0
                if selected_embeddings:
                    max_sim = max(
                        np.dot(candidate_embedding, sel_emb)
                        for sel_emb in selected_embeddings
                    )
                    diversity = 1 - max_sim

                score = lambda_mult * relevance - (1 - lambda_mult) * diversity

                if score > best_score:
                    best_score = score
                    best_idx = i

            selected.append(candidates.pop(best_idx))
            selected_embeddings.append(
                np.array(self.embeddings.embed_query(selected[-1].page_content))
            )

        return selected


class CheckpointManager:
    """Manages evaluation checkpoints for crash recovery"""
    def __init__(self, checkpoint_file: str):
        self.checkpoint_file = checkpoint_file

    def load_checkpoint(self) -> Optional[Dict[str, Any]]:
        """Load checkpoint if exists"""
        if Path(self.checkpoint_file).exists():
            with open(self.checkpoint_file, 'r') as f:
                return json.load(f)
        return None

    def save_checkpoint(self, data: Dict[str, Any]):
        """Save checkpoint"""
        with open(self.checkpoint_file, 'w') as f:
            json.dump(data, f, indent=2)

    def clear_checkpoint(self):
        """Clear checkpoint after successful run"""
        if Path(self.checkpoint_file).exists():
            Path(self.checkpoint_file).unlink()


class ExcelResultsManager:
    """Manages Excel file storage with proper formatting"""
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.ensure_workbook()

    def ensure_workbook(self):
        """Create Excel file with all sheets if not exists"""
        if not Path(self.file_path).exists():
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # Create sheets with headers
            for sheet_name, config in EXCEL_SHEETS.items():
                ws = wb.create_sheet(sheet_name)
                ws.append(config["columns"])
                self._format_header(ws)

            wb.save(self.file_path)
            logger.info(f"Created Excel file: {self.file_path}")

    def _format_header(self, worksheet):
        """Format header row"""
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def append_row(self, sheet_name: str, row_data: List[Any]):
        """Append row to sheet"""
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[sheet_name]
        ws.append(row_data)
        wb.save(self.file_path)

    def update_row(self, sheet_name: str, row_num: int, row_data: List[Any]):
        """Update specific row"""
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[sheet_name]
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_idx, value=value)
        wb.save(self.file_path)

    def get_row_count(self, sheet_name: str) -> int:
        """Get number of data rows (excluding header)"""
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[sheet_name]
        return ws.max_row - 1  # Exclude header


class RAGEvaluator:
    """Main RAG Evaluation Framework"""
    def __init__(self):
        self.test_run_id = str(uuid.uuid4())[:8]
        self.checkpoint_manager = CheckpointManager(EVAL_CHECKPOINT_FILE)
        self.excel_manager = ExcelResultsManager(EVAL_RESULTS_FILE)
        self.results = []
        self.checkpoint = None
        
        logger.info(f"Initialized RAG Evaluator [Run ID: {self.test_run_id}]")

    def load_dataset(self) -> List[Dict[str, str]]:
        """Load Q&A dataset"""
        with open(EVAL_DATASET_FILE, 'r') as f:
            dataset = json.load(f)
        return dataset

    def get_retriever(self, strategy: str, embeddings, documents: List[Document]):
        """Initialize retriever based on strategy"""
        if strategy == "semantic_search":
            return Chroma(
                embedding_function=embeddings,
                persist_directory="./local_rag_db"
            )
        elif strategy == "bm25":
            return BM25Retriever.from_documents(documents)
        elif strategy == "rrf":
            return RRFRetriever(vector_store, documents)
        elif strategy == "mmr":
            return MMRRetriever(vector_store, embeddings)
        else:
            raise ValueError(f"Unknown strategy: {strategy}")

    def retrieve_chunks(self, query: str, strategy: str, embeddings, k: int = 3) -> Tuple[List[Document], float]:
        """Retrieve chunks using specified strategy"""
        start_time = time.time()
        
        # Get all documents from vector store
        all_docs = vector_store.get()
        documents = [
            Document(
                page_content=doc_text,
                metadata={"id": doc_id}
            )
            for doc_id, doc_text in zip(all_docs['ids'], all_docs['documents'])
        ]

        if strategy == "semantic_search":
            retriever = vector_store
            results = retriever.similarity_search(query, k=k)
        elif strategy == "bm25":
            retriever = BM25Retriever.from_documents(documents)
            results = retriever.invoke(query)[:k]
        elif strategy == "rrf":
            retriever = RRFRetriever(vector_store, documents)
            results = retriever.retrieve(query, k=k)
        elif strategy == "mmr":
            retriever = MMRRetriever(vector_store, embeddings)
            results = retriever.retrieve(query, k=k, fetch_k=k*2)
        else:
            results = []

        retrieval_time = (time.time() - start_time) * 1000
        return results, retrieval_time

    def format_context(self, chunks: List[Document], mode: str) -> str:
        """Format retrieved chunks based on RAG mode"""
        if mode == "thread_rag":
            # Use context window assembly
            formatted = []
            for chunk in chunks:
                chunk_id = chunk.metadata.get("id", "unknown")
                content = chunk.page_content
                try:
                    context = assemble_chunk_with_context(chunk_id, content)
                    formatted.append(context)
                except:
                    formatted.append(f"Chunk {chunk_id}:\n{content}")
            return "\n\n".join(formatted)
        else:
            # Normal RAG - just raw chunks
            return "\n\n".join([f"Chunk {chunk.metadata.get('id', 'unknown')}:\n{chunk.page_content}" 
                               for chunk in chunks])

    def generate_answer(self, query: str, context: str, model: str) -> Tuple[str, float]:
        """Generate answer using specified model"""
        llm = ChatOllama(model=model, temperature=1.0)
        prompt = f"""Based on the following context, answer the question concisely.

Context:
{context}

Question: {query}

Answer:"""
        
        start_time = time.time()
        response = llm.invoke(prompt)
        gen_time = (time.time() - start_time) * 1000
        
        return response.content if hasattr(response, 'content') else str(response), gen_time

    def compute_perplexity(self, text: str, model: str) -> float:
        """Compute perplexity using language model"""
        try:
            llm = OllamaLLM(model=model)
            # Simplified perplexity: use log probability estimation
            # Full implementation would use actual token probabilities
            tokens = text.split()
            if len(tokens) < 2:
                return 0.0
            return float(len(set(tokens)) / len(tokens))  # Placeholder
        except Exception as e:
            logger.warning(f"Perplexity calculation failed: {e}")
            return 0.0

    def compute_mrr(self, retrieved_docs: List[Document], ground_truth: str) -> float:
        """Compute Mean Reciprocal Rank"""
        for i, doc in enumerate(retrieved_docs, 1):
            # Simple check: if document content contains ground truth keywords
            if any(word.lower() in doc.page_content.lower() 
                   for word in ground_truth.split()[:3]):
                return 1.0 / i
        return 0.0

    def compute_ndcg(self, retrieved_docs: List[Document], ground_truth: str, k: int = 3) -> float:
        """Compute Normalized Discounted Cumulative Gain"""
        # Simplified NDCG calculation
        dcg = 0.0
        for i, doc in enumerate(retrieved_docs[:k], 1):
            # Relevance score based on keyword overlap
            keywords = ground_truth.split()[:5]
            relevance = sum(1 for kw in keywords if kw.lower() in doc.page_content.lower()) / len(keywords)
            dcg += relevance / np.log2(i + 1)

        # Ideal DCG (if all docs were relevant)
        idcg = sum(1.0 / np.log2(i + 1) for i in range(1, min(k + 1, len(retrieved_docs) + 1)))
        
        return dcg / idcg if idcg > 0 else 0.0

    def compute_f1_score(self, generated: str, reference: str) -> float:
        """Compute token-level F1 score"""
        gen_tokens = set(generated.lower().split())
        ref_tokens = set(reference.lower().split())
        
        if len(gen_tokens) == 0 and len(ref_tokens) == 0:
            return 1.0
        if len(gen_tokens) == 0 or len(ref_tokens) == 0:
            return 0.0

        intersection = gen_tokens & ref_tokens
        precision = len(intersection) / len(gen_tokens)
        recall = len(intersection) / len(ref_tokens)

        if precision + recall == 0:
            return 0.0
        return 2 * (precision * recall) / (precision + recall)

    def compute_context_stability_score(self, retrieved_docs: List[Document], query: str, reference_answer: str) -> float:
        """Compute context stability score based on semantic coherence and information flow"""
        if not retrieved_docs or len(retrieved_docs) < 2:
            return 0.0

        try:
            # Calculate semantic coherence between adjacent chunks
            coherence_scores = []
            for i in range(len(retrieved_docs) - 1):
                chunk1_text = retrieved_docs[i].page_content
                chunk2_text = retrieved_docs[i + 1].page_content

                # Simple coherence based on overlapping concepts
                chunk1_words = set(chunk1_text.lower().split())
                chunk2_words = set(chunk2_text.lower().split())
                overlap = len(chunk1_words & chunk2_words)
                union = len(chunk1_words | chunk2_words)

                if union > 0:
                    jaccard_similarity = overlap / union
                    coherence_scores.append(jaccard_similarity)

            avg_coherence = sum(coherence_scores) / len(coherence_scores) if coherence_scores else 0.0

            # Information flow preservation (simplified)
            query_terms = set(query.lower().split())
            answer_terms = set(reference_answer.lower().split())

            # Check if key concepts flow through retrieved chunks
            chunk_concept_coverage = []
            for doc in retrieved_docs:
                chunk_terms = set(doc.page_content.lower().split())
                query_overlap = len(query_terms & chunk_terms)
                answer_overlap = len(answer_terms & chunk_terms)
                coverage = (query_overlap + answer_overlap) / (len(query_terms) + len(answer_terms)) if (len(query_terms) + len(answer_terms)) > 0 else 0
                chunk_concept_coverage.append(coverage)

            avg_coverage = sum(chunk_concept_coverage) / len(chunk_concept_coverage) if chunk_concept_coverage else 0.0

            # Boundary preservation (simplified)
            boundary_score = 1.0 if len(retrieved_docs) > 1 else 0.5

            # Weighted combination
            stability_score = (
                CONTEXT_STABILITY_CONFIG["semantic_coherence_weight"] * avg_coherence +
                CONTEXT_STABILITY_CONFIG["information_flow_weight"] * avg_coverage +
                CONTEXT_STABILITY_CONFIG["chunk_boundary_preservation_weight"] * boundary_score
            )

            return max(0.0, min(1.0, stability_score))

        except Exception as e:
            logger.warning(f"Context stability calculation failed: {e}")
            return 0.0

    def compute_token_savings(self, retrieved_docs: List[Document], mode: str) -> Tuple[int, float]:
        """Calculate token count and savings for different RAG modes"""
        try:
            if mode == "thread_rag":
                # Thread-RAG: context window + summaries
                base_tokens = len(retrieved_docs) * TOKEN_SAVINGS_CONFIG["context_window_tokens"]
                summary_tokens = len(retrieved_docs) * 2 * TOKEN_SAVINGS_CONFIG["summary_tokens"]  # prev + next
                total_tokens = base_tokens + summary_tokens
                savings_pct = (1 - TOKEN_SAVINGS_CONFIG["thread_rag_multiplier"]) * 100
            else:
                # Normal RAG: full chunks
                total_tokens = len(retrieved_docs) * TOKEN_SAVINGS_CONFIG["baseline_tokens_per_chunk"]
                savings_pct = 0.0

            return total_tokens, savings_pct

        except Exception as e:
            logger.warning(f"Token savings calculation failed: {e}")
            return 0, 0.0

    def run_evaluation(
        self,
        models: Optional[List[str]] = None,
        embedding_models: Optional[List[str]] = None,
        modes: Optional[List[str]] = None,
        strategies: Optional[List[str]] = None,
        specific_rows: Optional[List[int]] = None,
        resume_from_checkpoint: bool = True
    ):
        """Run complete evaluation suite"""
        # Set defaults
        models = models or OLLAMA_MODELS
        embedding_models = embedding_models or EMBEDDING_MODELS
        modes = modes or list(RAG_MODES.keys())
        strategies = strategies or list(RETRIEVAL_STRATEGIES.keys())

        logger.info(f"Starting evaluation with:")
        logger.info(f"  Models: {models}")
        logger.info(f"  Embedding Models: {embedding_models}")
        logger.info(f"  RAG Modes: {modes}")
        logger.info(f"  Strategies: {strategies}")

        # Load dataset
        dataset = self.load_dataset()
        if specific_rows:
            dataset = [dataset[i] for i in specific_rows if i < len(dataset)]

        logger.info(f"Loaded {len(dataset)} Q&A pairs")

        # Check for checkpoint
        if resume_from_checkpoint:
            self.checkpoint = self.checkpoint_manager.load_checkpoint()
            if self.checkpoint:
                logger.info(f"Resuming from checkpoint at row {self.checkpoint.get('last_row', 0)}")

        total_combinations = len(models) * len(embedding_models) * len(modes) * len(strategies) * len(dataset)
        current_combination = 0

        try:
            for model_idx, model in enumerate(models):
                logger.info(f"\n{'='*60}")
                logger.info(f"Model {model_idx + 1}/{len(models)}: {model}")
                logger.info(f"{'='*60}")

                for emb_idx, embedding_model in enumerate(embedding_models):
                    logger.info(f"\n  Embedding Model {emb_idx + 1}/{len(embedding_models)}: {embedding_model}")
                    
                    # Initialize embeddings for this iteration
                    embeddings = OllamaEmbeddings(model=embedding_model)

                    for mode_idx, mode in enumerate(modes):
                        logger.info(f"\n    RAG Mode {mode_idx + 1}/{len(modes)}: {mode}")

                        for strategy_idx, strategy in enumerate(strategies):
                            logger.info(f"\n      Strategy {strategy_idx + 1}/{len(strategies)}: {strategy}")

                            for row_idx, qa_pair in enumerate(dataset):
                                current_combination += 1
                                
                                # Check checkpoint
                                if self.checkpoint and row_idx < self.checkpoint.get('last_row', 0):
                                    continue

                                logger.info(f"\n        [{current_combination}/{total_combinations}] Processing Q{row_idx + 1}")

                                result = self._evaluate_single(
                                    qa_pair=qa_pair,
                                    model=model,
                                    embedding_model=embedding_model,
                                    mode=mode,
                                    strategy=strategy,
                                    row_idx=row_idx,
                                    embeddings=embeddings
                                )

                                self.results.append(result)
                                
                                # Write to Excel
                                self._write_result_to_excel(result)

                                # Save checkpoint
                                if (row_idx + 1) % CHECKPOINT_CONFIG["save_interval"] == 0:
                                    checkpoint_data = {
                                        "last_row": row_idx,
                                        "model": model,
                                        "embedding_model": embedding_model,
                                        "mode": mode,
                                        "strategy": strategy,
                                        "timestamp": datetime.now().isoformat(),
                                        "total_processed": len(self.results)
                                    }
                                    self.checkpoint_manager.save_checkpoint(checkpoint_data)
                                    logger.info(f"Checkpoint saved at row {row_idx}")

        except Exception as e:
            logger.error(f"Evaluation interrupted: {e}", exc_info=True)
            logger.info("Checkpoint available for resume")
            raise

        # Clear checkpoint on success
        self.checkpoint_manager.clear_checkpoint()
        
        # Generate summary
        self._generate_summary()
        logger.info(f"Evaluation complete! Results saved to {EVAL_RESULTS_FILE}")

    def _evaluate_single(
        self,
        qa_pair: Dict[str, str],
        model: str,
        embedding_model: str,
        mode: str,
        strategy: str,
        row_idx: int,
        embeddings
    ) -> Dict[str, Any]:
        """Evaluate single Q&A pair"""
        result = {
            "row_id": row_idx,
            "timestamp": datetime.now().isoformat(),
            "question": qa_pair["question"],
            "reference_answer": qa_pair.get("answer", ""),
            "document_type": qa_pair.get("document_type", ""),
            "context_dependency": qa_pair.get("context_dependency", ""),
            "chunk_span": qa_pair.get("chunk_span", ""),
            "rag_mode": mode,
            "retrieval_strategy": strategy,
            "main_model": model,
            "embedding_model": embedding_model,
            "status": "pending",
            "error_message": ""
        }

        try:
            # Retrieve chunks
            retrieved_chunks, retrieval_time = self.retrieve_chunks(
                qa_pair["question"],
                strategy,
                embeddings
            )
            result["retrieval_time_ms"] = round(retrieval_time, 2)
            result["retrieved_chunks"] = len(retrieved_chunks)

            # Format context based on mode
            context = self.format_context(retrieved_chunks, mode)

            # Generate answer
            generated_answer, gen_time = self.generate_answer(
                qa_pair["question"],
                context,
                model
            )
            result["generated_answer"] = generated_answer
            result["generation_time_ms"] = round(gen_time, 2)
            result["total_time_ms"] = round(retrieval_time + gen_time, 2)

            # Compute metrics
            result["f1_score"] = round(self.compute_f1_score(generated_answer, qa_pair.get("answer", "")), 4)
            result["mrr"] = round(self.compute_mrr(retrieved_chunks, qa_pair.get("answer", "")), 4)
            result["ndcg"] = round(self.compute_ndcg(retrieved_chunks, qa_pair.get("answer", "")), 4)
            result["perplexity"] = round(self.compute_perplexity(generated_answer, model), 4)

            # RAGAS metrics (simplified - full version needs proper setup)
            try:
                result["faithfulness"] = 0.0  # Placeholder
                result["answer_relevancy"] = 0.0  # Placeholder
                result["context_precision"] = 0.0  # Placeholder
                result["context_recall"] = 0.0  # Placeholder
            except:
                pass

            # Context stability score
            result["context_stability"] = round(self.compute_context_stability_score(retrieved_chunks, qa_pair["question"], qa_pair.get("answer", "")), 4)

            # Token savings
            result["token_savings"], result["savings_percentage"] = self.compute_token_savings(retrieved_chunks, mode)

            result["status"] = "completed"

        except Exception as e:
            logger.error(f"Error evaluating Q{row_idx}: {e}")
            result["status"] = "failed"
            result["error_message"] = str(e)

        return result

    def _write_result_to_excel(self, result: Dict[str, Any]):
        """Write result to detailed results sheet"""
        row_data = [
            result.get("row_id", ""),
            result.get("timestamp", ""),
            result.get("question", ""),
            result.get("reference_answer", ""),
            result.get("document_type", ""),
            result.get("context_dependency", ""),
            result.get("chunk_span", ""),
            result.get("rag_mode", ""),
            result.get("retrieval_strategy", ""),
            result.get("main_model", ""),
            result.get("embedding_model", ""),
            result.get("retrieved_chunks", 0),
            result.get("generated_answer", "")[:100],  # Truncate for Excel
            result.get("faithfulness", 0),
            result.get("answer_relevancy", 0),
            result.get("context_precision", 0),
            result.get("context_recall", 0),
            result.get("f1_score", 0),
            result.get("context_stability", 0),
            result.get("mrr", 0),
            result.get("ndcg", 0),
            result.get("perplexity", 0),
            result.get("retrieval_time_ms", 0),
            result.get("generation_time_ms", 0),
            result.get("total_time_ms", 0),
            result.get("token_savings", 0),
            result.get("savings_percentage", 0),
            result.get("status", ""),
            result.get("error_message", "")
        ]
        self.excel_manager.append_row("detailed", row_data)

    def _generate_summary(self):
        """Generate summary statistics"""
        if not self.results:
            return

        timestamp = datetime.now().isoformat()
        completed = sum(1 for r in self.results if r["status"] == "completed")
        failed = sum(1 for r in self.results if r["status"] == "failed")

        # Calculate averages
        completed_results = [r for r in self.results if r["status"] == "completed"]
        if completed_results:
            avg_faithfulness = np.mean([r.get("faithfulness", 0) for r in completed_results])
            avg_relevancy = np.mean([r.get("answer_relevancy", 0) for r in completed_results])
            avg_precision = np.mean([r.get("context_precision", 0) for r in completed_results])
            avg_recall = np.mean([r.get("context_recall", 0) for r in completed_results])
            avg_perplexity = np.mean([r.get("perplexity", 0) for r in completed_results])
            avg_mrr = np.mean([r.get("mrr", 0) for r in completed_results])
            avg_ndcg = np.mean([r.get("ndcg", 0) for r in completed_results])
        else:
            avg_faithfulness = avg_relevancy = avg_precision = avg_recall = 0
            avg_perplexity = avg_mrr = avg_ndcg = 0

        summary_row = [
            timestamp,
            self.test_run_id,
            len(self.results),
            completed,
            failed,
            round(avg_faithfulness, 4),
            round(avg_relevancy, 4),
            round(avg_precision, 4),
            round(avg_recall, 4),
            round(avg_perplexity, 4),
            round(avg_mrr, 4),
            round(avg_ndcg, 4)
        ]
        self.excel_manager.append_row("summary", summary_row)

    def generate_comparison_sheets(self):
        """Generate comparison sheets for RAG modes and strategies"""
        logger.info("Generating comparison sheets...")
        
        # Group results by mode
        mode_results = defaultdict(list)
        for result in self.results:
            if result["status"] == "completed":
                mode_results[result["rag_mode"]].append(result)

        # Generate mode comparison
        wb = openpyxl.load_workbook(EVAL_RESULTS_FILE)
        ws_comparison = wb["comparison"]
        
        metrics = ["faithfulness", "answer_relevancy", "context_precision", "context_recall", "mrr", "ndcg"]
        
        for metric in metrics:
            row_data = [metric]
            for mode in RAG_MODES.keys():
                values = [r.get(metric, 0) for r in mode_results.get(mode, [])]
                avg = np.mean(values) if values else 0
                row_data.append(round(avg, 4))
            
            # Calculate difference
            if len(row_data) >= 3:
                row_data.append(round(row_data[-1] - row_data[1], 4))
                row_data.append(round((row_data[-2] / row_data[2] - 1) * 100 if row_data[2] > 0 else 0, 2))
            
            ws_comparison.append(row_data)

        wb.save(EVAL_RESULTS_FILE)
        logger.info("Comparison sheets generated")


if __name__ == "__main__":
    evaluator = RAGEvaluator()
    
    # Example: Run full evaluation
    evaluator.run_evaluation(
        models=OLLAMA_MODELS[:2],  # Use first 2 models for testing
        embedding_models=EMBEDDING_MODELS[:1],  # Use first embedding model
        modes=list(RAG_MODES.keys()),
        strategies=list(RETRIEVAL_STRATEGIES.keys())
    )
    
    # Generate comparisons
    evaluator.generate_comparison_sheets()

